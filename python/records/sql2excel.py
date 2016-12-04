# -*- coding: utf-8 -*-
import pymysql as mc  
from openpyxl.writer.excel import ExcelWriter  
from openpyxl.workbook import Workbook  
#it's different between workbook and worksheet  
conn = mc.connect(user='root', password='', host='127.0.0.1', database='maolizi',charset="utf8")  
cur = conn.cursor()  
  
# query language  
sql = "select name,age from user"  
data = cur.execute(sql)  
index = cur.description 

# use fetchall method to get a list, some tuples in it  
data_list = cur.fetchall() 
cur.close()  
conn.close()
# print (data_list)  
  
wb = Workbook()  
ws = wb.worksheets[0]  
ws.title = 'data'  
  
file_name = r'user.xlsx'  
rows = len(data_list)  
cols = len(data_list[0])  
  
# same as read data  
for rx in range(rows):  
	for cx in range(cols): 
		if rx==0:
			ws.cell(row=1,column=cx+1).value=index[cx][0]
		ws.cell(row=rx+2, column=cx+1).value = data_list[rx][cx]  
  
# use ExcelWriter method to write file and save it  
wb.save('user.xlsx') 
